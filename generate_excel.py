import json
import os
import re
import sys

# Auto-install openpyxl if not present
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

workspace_dir = os.path.dirname(os.path.abspath(__file__))
product_cache_path = os.path.join(workspace_dir, "product_cache.json")
product_js_path = os.path.join(workspace_dir, "product.js")
index_html_path = os.path.join(workspace_dir, "index.html")
gitignore_path = os.path.join(workspace_dir, ".gitignore")
excel_output_path = os.path.join(workspace_dir, "noku_products.xlsx")

# 1. Parse index.html product cards (Primary dimensions source for cards in index)
index_dims = {}
index_prices = {}

if os.path.exists(index_html_path):
    with open(index_html_path, 'r', encoding='utf-8') as f:
        index_html = f.read()
    
    # Extract data-handle and data-dim
    dim_matches = re.findall(r'class="product-card"[^>]*data-handle="([^"]+)"[^>]*data-dim="([^"]+)"', index_html)
    for handle, dim in dim_matches:
        index_dims[handle] = dim.strip()

    # Also search for price inside card
    # Find all product-card elements and extract handle, name, price
    # Simple regex to scan product-card blocks
    card_blocks = re.findall(r'class="product-card"[^>]*data-handle="([^"]+)".*?class="product-price">₹([^<]+)<', index_html, re.DOTALL)
    for handle, price_str in card_blocks:
        price_cleaned = price_str.replace(',', '').strip()
        index_prices[handle] = float(price_cleaned)

# 2. Parse product.js dimensions (Secondary dimensions source)
js_dims = {}
if os.path.exists(product_js_path):
    with open(product_js_path, 'r', encoding='utf-8') as f:
        product_js = f.read()
    handle_blocks = re.findall(r'"([^"]+)":\s*\{[^}]*?dimension:\s*"([^"]+)"', product_js, re.DOTALL)
    for handle, dim in handle_blocks:
        js_dims[handle] = dim.strip()

# 3. Hardcoded lookup from searchIndex.json for products not fully specified in product.js
search_index_dims = {
    "round-dining-chair": "50cm x 50cm x 77cm (Seat Height: 45cm)",
    "round-dining-table": "183cm x 107cm x 76cm",
    "poster-bed": "204cm x 220cm x 197cm (Seat Height: 48cm)",
    "dining-table": "198cm x 92cm x 75cm / 183cm x 92cm x 75cm",
    "study-chair": "60cm x 60cm x 84cm (Seat Height: 45cm)",
    "barstool-01": "48cm x 50cm x 100cm (Seat Height: 75cm)"
}

# Combine dimensions mapping (HTML > JS > searchIndex)
def get_dimension(handle):
    return index_dims.get(handle) or js_dims.get(handle) or search_index_dims.get(handle) or "Not specified"

# 4. Load product_cache.json
if os.path.exists(product_cache_path):
    with open(product_cache_path, 'r', encoding='utf-8') as f:
        product_cache = json.load(f)
else:
    product_cache = {}

# Compile all products and variants
all_products = []
for gid, p in product_cache.items():
    title = p.get('title')
    handle = p.get('handle')
    variants = p.get('variants', {})
    
    dim = get_dimension(handle)
    
    # Get variant prices
    v_list = []
    for v_gid, v in variants.items():
        v_title = v.get('title')
        v_price = v.get('price')
        v_list.append({
            "variant_id": v_gid,
            "title": v_title,
            "price": v_price
        })
    
    all_products.append({
        "gid": gid,
        "title": title,
        "handle": handle,
        "dimension": dim,
        "variants": v_list
    })

# Sort products by title
all_products.sort(key=lambda x: x['title'])

# 5. Create Excel Workbook
wb = openpyxl.Workbook()

# Sheet 1: Product Summary
ws_summary = wb.active
ws_summary.title = "Products Summary"

# Styling definitions
header_fill = PatternFill(start_color="1A3636", end_color="1A3636", fill_type="solid") # Dark elegant Slate green/grey
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=14, bold=True, color="1A3636")
bold_font = Font(name="Calibri", size=11, bold=True)
regular_font = Font(name="Calibri", size=11)
italic_font = Font(name="Calibri", size=9, italic=True)

thin_side = Side(border_style="thin", color="D3D3D3")
thick_bottom = Side(border_style="medium", color="1A3636")
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)

# Add sheet title block
ws_summary.append(["NOKU STUDIO — PRODUCT PORTFOLIO"])
ws_summary.cell(row=1, column=1).font = title_font
ws_summary.row_dimensions[1].height = 25
ws_summary.append(["Generated from workspace cache files. All prices in INR."])
ws_summary.cell(row=2, column=1).font = italic_font
ws_summary.append([]) # Empty row

# Headers for Sheet 1
summary_headers = ["Product Name", "Dimensions", "Price Range", "Number of Variants"]
ws_summary.append(summary_headers)
header_row_idx = 4
ws_summary.row_dimensions[header_row_idx].height = 24

for col_idx, header in enumerate(summary_headers, start=1):
    cell = ws_summary.cell(row=header_row_idx, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left" if col_idx < 3 else "right" if col_idx == 4 else "center", vertical="center")
    cell.border = border_header

# Populate Summary Sheet
row_idx = 5
for p in all_products:
    title = p['title']
    dim = p['dimension']
    variants = p['variants']
    
    # Calculate price range
    prices = [v['price'] for v in variants if v['price'] is not None]
    if prices:
        min_p = min(prices)
        max_p = max(prices)
        if min_p == max_p:
            if min_p == 0:
                price_range = "TBD"
            else:
                price_range = f"₹{min_p:,.0f}"
        else:
            price_range = f"₹{min_p:,.0f} - ₹{max_p:,.0f}"
    else:
        # Fallback to index.html price if available
        index_p = index_prices.get(p['handle'])
        if index_p:
            price_range = f"₹{index_p:,.0f}"
        else:
            price_range = "Not specified"
            
    num_variants = len(variants)
    
    ws_summary.append([title, dim, price_range, num_variants])
    
    # Style the data row
    ws_summary.row_dimensions[row_idx].height = 20
    for col_idx in range(1, 5):
        cell = ws_summary.cell(row=row_idx, column=col_idx)
        cell.font = regular_font
        cell.border = border_all
        if col_idx == 3:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 4:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
    row_idx += 1


# Sheet 2: Product Variants Breakout
ws_variants = wb.create_sheet(title="Product Variants")

# Add sheet title block
ws_variants.append(["NOKU STUDIO — DETAILED PRODUCT VARIANTS"])
ws_variants.cell(row=1, column=1).font = title_font
ws_variants.row_dimensions[1].height = 25
ws_variants.append(["Breakdown of every available variant from store cache. All prices in INR."])
ws_variants.cell(row=2, column=1).font = italic_font
ws_variants.append([]) # Empty row

# Headers for Sheet 2
variant_headers = ["Product Name", "Variant Name", "Dimensions", "Price"]
ws_variants.append(variant_headers)
v_header_row_idx = 4
ws_variants.row_dimensions[v_header_row_idx].height = 24

for col_idx, header in enumerate(variant_headers, start=1):
    cell = ws_variants.cell(row=v_header_row_idx, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left" if col_idx != 4 else "right", vertical="center")
    cell.border = border_header

# Populate Variant Sheet
v_row_idx = 5
for p in all_products:
    title = p['title']
    dim = p['dimension']
    variants = p['variants']
    
    if not variants:
        # If no variants, put a single row with TBD price
        ws_variants.append([title, "Base Style", dim, 0])
        ws_variants.row_dimensions[v_row_idx].height = 18
        for col_idx in range(1, 5):
            cell = ws_variants.cell(row=v_row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = border_all
            if col_idx == 4:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '"₹"#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        v_row_idx += 1
    else:
        for v in variants:
            v_title = v['title']
            v_price = v['price']
            
            ws_variants.append([title, v_title, dim, v_price])
            ws_variants.row_dimensions[v_row_idx].height = 18
            for col_idx in range(1, 5):
                cell = ws_variants.cell(row=v_row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = border_all
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if v_price == 0:
                        cell.value = "TBD"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.number_format = '"₹"#,##0'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            v_row_idx += 1

# Auto-adjust column widths for both sheets to make it look clean
for ws in [ws_summary, ws_variants]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't check title rows for length
        for cell in col:
            if cell.row > 3 and cell.value:
                # Format checks for currency vs regular text
                val_str = str(cell.value)
                if isinstance(cell.value, (int, float)) and cell.number_format and "₹" in cell.number_format:
                    val_str = f"₹{cell.value:,.0f}"
                max_len = max(max_len, len(val_str))
                
        # Give padding
        adjusted_width = max(max_len + 4, 12)
        # Limit width for dimension or title column so it doesn't extend too far
        if col_letter in ['A', 'B', 'C']:
            adjusted_width = min(adjusted_width, 45)
            
        ws.column_dimensions[col_letter].width = adjusted_width

# Save Workbook
wb.save(excel_output_path)
print(f"Excel file created successfully at: {excel_output_path}")

# 6. Update .gitignore to exclude noku_products.xlsx
if os.path.exists(gitignore_path):
    with open(gitignore_path, 'r', encoding='utf-8') as f:
        gitignore_content = f.read()
    
    if "noku_products.xlsx" not in gitignore_content:
        # Append noku_products.xlsx to the gitignore content
        # Ensure a trailing newline
        new_content = gitignore_content
        if not gitignore_content.endswith('\n'):
            new_content += '\n'
        new_content += "# Generated Product Excel Spreadsheet\nnoku_products.xlsx\n"
        
        with open(gitignore_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        print(".gitignore updated to exclude noku_products.xlsx")
    else:
        print("noku_products.xlsx is already excluded in .gitignore")
else:
    # Create a gitignore
    with open(gitignore_path, 'w', encoding='utf-8') as f:
        f.write("# Generated Product Excel Spreadsheet\nnoku_products.xlsx\n")
    print(".gitignore created and noku_products.xlsx excluded")
