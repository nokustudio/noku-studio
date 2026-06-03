"""
update_display_items.py
=======================
Noku Studio — Product Display Status Sync Script

Reads noku_products.xlsx and updates the three JavaScript files
(product.js, products.js, shopify-integration.js) to reflect:
  1. Which products are "Display" vs "Sale"
  2. New products added to the Excel sheet

It updates the following arrays in each JS file:
  - DISPLAY_ONLY_HANDLES         (top-level const)
  - displayTitles                (inside isDisplayItem())
  - displayVariantIds            (inside isDisplayItem())
  - displayHandles + displayTitles + displayVariantIds
                                 (inside updateCartUI() checkout filter)

Usage:
    python update_display_items.py

Requirements:
    pip install openpyxl
"""

import json
import os
import re
import sys
from datetime import datetime

# Force UTF-8 output on Windows to avoid cp1252 encoding errors
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ─── AUTO-INSTALL openpyxl IF MISSING ───────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ─── FILE PATHS ─────────────────────────────────────────────────────────────
WORKSPACE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH        = os.path.join(WORKSPACE, "noku_products.xlsx")
CACHE_PATH        = os.path.join(WORKSPACE, "product_cache.json")
PRODUCT_JS        = os.path.join(WORKSPACE, "product.js")
PRODUCTS_JS       = os.path.join(WORKSPACE, "products.js")
SHOPIFY_JS        = os.path.join(WORKSPACE, "shopify-integration.js")


# ─── STEP 1: LOAD EXCEL SHEET ───────────────────────────────────────────────
def load_display_products_from_excel(excel_path):
    """
    Reads Products Summary sheet.
    Returns a dict: { 'Product Name (lower)': 'Display' | 'Sale' }
    """
    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel file not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)

    if "Products Summary" not in wb.sheetnames:
        print("[ERROR] Sheet 'Products Summary' not found in the Excel file.")
        sys.exit(1)

    sh = wb["Products Summary"]

    # Expected columns (row 4 is header):
    #   Col 1: Product Name   Col 5: Purchase mode
    products = {}
    for row in sh.iter_rows(min_row=5):  # data starts at row 5
        name_cell = row[0].value
        mode_cell = row[4].value if len(row) >= 5 else None

        if not name_cell or not mode_cell:
            continue

        name = str(name_cell).strip()
        mode = str(mode_cell).strip()
        products[name] = mode

    return products


# ─── STEP 2: LOAD PRODUCT CACHE ─────────────────────────────────────────────
def load_product_cache(cache_path):
    """
    Returns the product_cache.json as a dict keyed by GID.
    """
    if not os.path.exists(cache_path):
        print(f"[WARNING] product_cache.json not found at {cache_path}. Variant IDs cannot be resolved.")
        return {}

    with open(cache_path, "r", encoding="utf-8") as f:
        return json.load(f)


# ─── STEP 3: MAP EXCEL TITLES → HANDLES + VARIANT IDS ──────────────────────
def resolve_display_data(excel_products, product_cache):
    """
    For every product marked as 'Display' in Excel:
      - Finds its handle in the product cache (by title match)
      - Collects all variant IDs (numeric suffix from GID)

    Returns:
      display_handles   : list of handles        e.g. ['dining-table', 'sofa-2', ...]
      display_titles    : list of titles (lower)  e.g. ['dining table', 'grooved sofa', ...]
      display_variant_ids: list of variant ID strings
      sale_handles      : list of handles for Sale products (for reporting)
      unmatched_names   : names in Excel that couldn't be matched to cache
    """
    # Build a lookup: lowercase title -> (gid, product_data)
    title_to_product = {}
    for gid, pdata in product_cache.items():
        title = pdata.get("title", "").strip().lower()
        title_to_product[title] = (gid, pdata)

    display_handles     = []
    display_titles      = []
    display_variant_ids = []
    sale_handles        = []
    unmatched_names     = []

    for name, mode in excel_products.items():
        name_lower = name.strip().lower()

        if name_lower not in title_to_product:
            unmatched_names.append(name)
            continue

        gid, pdata = title_to_product[name_lower]
        handle = pdata.get("handle", "")
        variants = pdata.get("variants", {})

        if mode.strip().lower() == "display":
            if handle and handle not in display_handles:
                display_handles.append(handle)

            # Build display title entries: both raw title and common variations
            raw_title = name_lower
            if raw_title not in display_titles:
                display_titles.append(raw_title)
            # Also add hyphenated form if title has spaces
            hyphenated = raw_title.replace(" ", "-")
            if hyphenated != raw_title and hyphenated not in display_titles:
                display_titles.append(hyphenated)

            # Collect numeric part of each variant GID
            for v_gid in variants.keys():
                # GID format: "gid://shopify/ProductVariant/12345678"
                numeric_id = v_gid.split("/")[-1]
                if numeric_id and numeric_id not in display_variant_ids:
                    display_variant_ids.append(numeric_id)
        else:
            if handle and handle not in sale_handles:
                sale_handles.append(handle)

    return display_handles, display_titles, display_variant_ids, sale_handles, unmatched_names


# ─── STEP 4: FORMATTERS ─────────────────────────────────────────────────────
def format_handles_array(handles, indent="  "):
    """Formats a list of handles as a JS array string (multi-line, single-quoted)."""
    if not handles:
        return "[]"
    items = [f"{indent}'{h}'" for h in handles]
    return "[\n" + ",\n".join(items) + "\n]"


def format_titles_array(titles, indent="    "):
    """Formats a list of titles as a JS array string (multi-line, double-quoted)."""
    if not titles:
        return "[]"
    items = [f'{indent}"{t}"' for t in titles]
    return "[\n" + ",\n".join(items) + "\n  ]"


def format_variant_ids_array(ids, indent="    ", chunk_size=10):
    """Formats variant IDs as a JS array string, chunked 10 per line."""
    if not ids:
        return "[]"
    lines = []
    for i in range(0, len(ids), chunk_size):
        chunk = ids[i:i + chunk_size]
        quoted = [f'"{vid}"' for vid in chunk]
        lines.append(f'{indent}' + ", ".join(quoted))
    return "[\n" + ",\n".join(lines) + "\n  ]"


# ─── STEP 5: PERFORM REGEX REPLACEMENTS ─────────────────────────────────────
def update_js_file(filepath, display_handles, display_titles, display_variant_ids):
    """
    Updates the given JS file in-place. Returns (changed: bool, changes_description: list).
    """
    if not os.path.exists(filepath):
        print(f"[ERROR] JS file not found: {filepath}")
        return False, []

    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    original_content = content
    changes = []

    # ── 5a. Replace top-level DISPLAY_ONLY_HANDLES ──────────────────────────
    new_handles_str = format_handles_array(display_handles)
    handles_pattern = re.compile(
        r"(// Display-only products list.*?\n)"
        r"(const DISPLAY_ONLY_HANDLES\s*=\s*)\[.*?\];",
        re.DOTALL
    )
    replacement = rf"\g<1>\g<2>{new_handles_str};"
    new_content, n_subs = re.subn(handles_pattern, replacement, content)
    if n_subs > 0:
        changes.append(f"  [OK] DISPLAY_ONLY_HANDLES updated ({len(display_handles)} handles, {n_subs} location(s))")
        content = new_content
    else:
        changes.append("  [!!] DISPLAY_ONLY_HANDLES pattern not found - skipped")

    # ── 5b. Replace displayTitles inside isDisplayItem() ────────────────────
    new_titles_str = format_titles_array(display_titles)
    titles_pattern = re.compile(
        r"(const displayTitles\s*=\s*)\[.*?\];",
        re.DOTALL
    )
    new_content, n_subs = re.subn(titles_pattern, rf"\g<1>{new_titles_str};", content)
    if n_subs > 0:
        changes.append(f"  [OK] displayTitles updated ({len(display_titles)} titles, {n_subs} location(s))")
        content = new_content
    else:
        changes.append("  [!!] displayTitles pattern not found - skipped")

    # ── 5c. Replace displayVariantIds (all occurrences) ─────────────────────
    new_vids_str = format_variant_ids_array(display_variant_ids)
    vids_pattern = re.compile(
        r"(const displayVariantIds\s*=\s*)\[.*?\];",
        re.DOTALL
    )
    new_content, n_subs = re.subn(vids_pattern, rf"\g<1>{new_vids_str};", content)
    if n_subs > 0:
        changes.append(f"  [OK] displayVariantIds updated ({len(display_variant_ids)} IDs, {n_subs} location(s))")
        content = new_content
    else:
        changes.append("  [!!] displayVariantIds pattern not found - skipped")

    # ── 5d. Replace displayHandles inside updateCartUI() ────────────────────
    new_dh_str = format_handles_array(display_handles, indent="    ")
    dh_pattern = re.compile(
        r"(const displayHandles\s*=\s*)\[.*?\];",
        re.DOTALL
    )
    new_content, n_subs = re.subn(dh_pattern, rf"\g<1>{new_dh_str};", content)
    if n_subs > 0:
        changes.append(f"  [OK] displayHandles (in updateCartUI) updated ({n_subs} location(s))")
        content = new_content
    else:
        changes.append("  [!!] displayHandles pattern not found - skipped")

    changed = content != original_content
    if changed:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)

    return changed, changes


# ─── STEP 6: CHECK FOR NEW PRODUCTS IN EXCEL NOT YET IN CACHE ───────────────
def check_new_products(excel_products, product_cache):
    """
    Reports any product in the Excel that has no matching title in the cache.
    These are likely new products that need to be added to Shopify and product_cache.json.
    """
    title_to_product = {
        pdata.get("title", "").strip().lower(): pdata
        for pdata in product_cache.values()
    }
    new_products = []
    for name in excel_products.keys():
        if name.strip().lower() not in title_to_product:
            new_products.append(name)
    return new_products


# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Noku Studio — Display/Sale Sync Script")
    print(f"  Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Load data
    print("\n[1/4] Reading Excel file...")
    excel_products = load_display_products_from_excel(EXCEL_PATH)
    print(f"      Found {len(excel_products)} products in 'Products Summary':")
    for name, mode in sorted(excel_products.items()):
        icon = "[DISPLAY]" if mode.lower() == "display" else "[SALE]"
        print(f"      {icon} {name} - {mode}")

    print("\n[2/4] Loading product_cache.json...")
    product_cache = load_product_cache(CACHE_PATH)
    print(f"      {len(product_cache)} products loaded from cache.")

    print("\n[3/4] Resolving display handles, titles and variant IDs...")
    display_handles, display_titles, display_variant_ids, sale_handles, unmatched = \
        resolve_display_data(excel_products, product_cache)

    print(f"\n      Display-only handles ({len(display_handles)}):")
    for h in display_handles:
        print(f"        - {h}")

    print(f"\n      Display titles ({len(display_titles)}):")
    for t in display_titles:
        print(f"        - {t}")

    print(f"\n      Display variant IDs: {len(display_variant_ids)} total")

    if unmatched:
        print(f"\n      [!] {len(unmatched)} product(s) from Excel not found in cache:")
        for u in unmatched:
            print(f"        - {u}")
        print("        -> These may be new products not yet on Shopify / in product_cache.json.")
        print("           Run: node product-checker.js  to refresh the cache.")

    # Check for new products
    new_products = check_new_products(excel_products, product_cache)
    if new_products:
        print(f"\n      [NEW] New products detected (in Excel but not in cache):")
        for np in new_products:
            print(f"        - {np}")

    # Update JS files
    print("\n[4/4] Updating JavaScript files...")

    js_files = [
        ("product.js",               PRODUCT_JS),
        ("products.js",              PRODUCTS_JS),
        ("shopify-integration.js",   SHOPIFY_JS),
    ]

    any_updated = False
    for label, path in js_files:
        print(f"\n  → {label}")
        changed, changes = update_js_file(
            path, display_handles, display_titles, display_variant_ids
        )
        for c in changes:
            print(c)
        if changed:
            print(f"  [SAVED] File updated and saved.")
            any_updated = True
        else:
            print(f"  [OK] No changes needed.")

    print("\n" + "=" * 60)
    if any_updated:
        print("  [SUCCESS] Update complete! JS files have been synced with Excel.")
        print("\n  Next steps:")
        print("  1. Test locally: open product.html?handle=<display-handle>")
        print("     and verify the 'Get in Touch' button appears.")
        print("  2. Commit and push changes:")
        print("     git add product.js products.js shopify-integration.js")
        print("     git commit -m \"sync: update display-only product list from Excel\"")
        print("     git push")
    else:
        print("  [OK] All JS files are already up to date. No changes made.")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
