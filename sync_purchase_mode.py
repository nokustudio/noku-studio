"""
Noku Studio — Sync Excel "Purchase mode" from the Shopify metafield.

Source of truth is the Shopify product metafield custom.purchasemode (boolean):
  true  -> "Sale"     (buyable / Add to Cart)
  false -> "Display"  (inquire-only)

This rewrites column 5 ("Purchase mode") of the "Products Summary" sheet in
noku_products.xlsx in place, matching rows by product name. Rows with no live
match are left unchanged and reported. Run after changing purchasemode in
Shopify Admin so the spreadsheet reflects what the website actually does.
"""
import os, json, sys
import urllib.request
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

WORKSPACE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(WORKSPACE, "noku_products.xlsx")

SHOP = "6b5390-f8.myshopify.com"
API_VERSION = "2024-04"
TOKEN = "7b62ad5d7d665bebe383ff2d3c36c0b0"

QUERY = """
{
  products(first: 100) {
    edges { node {
      title
      purchaseMode: metafield(namespace: "custom", key: "purchasemode") { value }
    } }
  }
}
"""


def fetch_modes():
    """Return { product_title_lower: "Sale" | "Display" } from Shopify."""
    req = urllib.request.Request(
        f"https://{SHOP}/api/{API_VERSION}/graphql.json",
        data=json.dumps({"query": QUERY}).encode(),
        headers={
            "Content-Type": "application/json",
            "X-Shopify-Storefront-Access-Token": TOKEN,
        },
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.load(resp)
    if data.get("errors"):
        raise SystemExit(f"GraphQL errors: {data['errors']}")
    modes = {}
    for edge in data["data"]["products"]["edges"]:
        node = edge["node"]
        mf = node.get("purchaseMode") or {}
        buyable = mf.get("value") == "true"
        modes[node["title"].strip().lower()] = "Sale" if buyable else "Display"
    return modes


def main():
    if not os.path.exists(EXCEL_PATH):
        raise SystemExit(f"Excel not found: {EXCEL_PATH}")

    modes = fetch_modes()
    print(f"Fetched purchasemode for {len(modes)} live products.\n")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Products Summary"]

    changed, unmatched = [], []
    for r in range(5, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        key = str(name).strip().lower()
        if key not in modes:
            unmatched.append(str(name))
            continue
        new_mode = modes[key]
        cell = ws.cell(row=r, column=5)
        if cell.value != new_mode:
            changed.append(f"{name}: {cell.value!r} -> {new_mode!r}")
            cell.value = new_mode

    wb.save(EXCEL_PATH)

    print(f"Updated {len(changed)} row(s):")
    for c in changed:
        print("  " + c)
    if unmatched:
        print(f"\nNo live match (left unchanged): {', '.join(unmatched)}")
    print(f"\nSaved: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
