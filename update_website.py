import json
import os
import re
import sys
import urllib.request
import urllib.error

# Auto-install openpyxl if not present
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

# Define file paths relative to this script
WORKSPACE = os.path.dirname(os.path.abspath(__file__))
SHOPIFY_JS = os.path.join(WORKSPACE, "shopify-integration.js")
PRODUCTS_JS = os.path.join(WORKSPACE, "products.js")
PRODUCT_JS = os.path.join(WORKSPACE, "product.js")
INDEX_HTML = os.path.join(WORKSPACE, "index.html")
CACHE_PATH = os.path.join(WORKSPACE, "product_cache.json")
EXCEL_PATH = os.path.join(WORKSPACE, "noku_products.xlsx")
GITIGNORE_PATH = os.path.join(WORKSPACE, ".gitignore")


def load_shopify_config():
    if not os.path.exists(SHOPIFY_JS):
        print(f"[ERROR] shopify-integration.js not found at {SHOPIFY_JS}")
        sys.exit(1)
        
    with open(SHOPIFY_JS, "r", encoding="utf-8") as f:
        content = f.read()
        
    token_match = re.search(r"storefrontAccessToken:\s*'([^']+)'", content)
    domain_match = re.search(r"shopDomain:\s*'([^']+)'", content)
    version_match = re.search(r"apiVersion:\s*'([^']+)'", content)
    
    if not token_match or not domain_match:
        print("[ERROR] Could not parse Shopify credentials from shopify-integration.js")
        sys.exit(1)
        
    return {
        "storefrontAccessToken": token_match.group(1),
        "shopDomain": domain_match.group(1),
        "apiVersion": version_match.group(1) if version_match else "2024-04"
    }


def fetch_products(config):
    url = f"https://{config['shopDomain']}/api/{config['apiVersion']}/graphql.json"
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Storefront-Access-Token": config["storefrontAccessToken"]
    }
    
    # Storefront GraphQL Query to fetch products and variants
    query = """
    query getAllProducts {
      products(first: 100) {
        edges {
          node {
            id
            title
            handle
            featuredImage {
              url
            }
            variants(first: 100) {
              edges {
                node {
                  id
                  title
                  price {
                    amount
                    currencyCode
                  }
                }
              }
            }
          }
        }
      }
    }
    """
    
    req_body = json.dumps({"query": query}).encode("utf-8")
    req = urllib.request.Request(url, data=req_body, headers=headers, method="POST")
    
    try:
        print(f"Connecting to Shopify storefront: {config['shopDomain']}...")
        with urllib.request.urlopen(req) as response:
            res_body = response.read().decode("utf-8")
            result = json.loads(res_body)
            
            if "errors" in result:
                print(f"[ERROR] GraphQL errors: {result['errors']}")
                sys.exit(1)
                
            return result.get("data", {}).get("products", {}).get("edges", [])
    except urllib.error.URLError as e:
        print(f"[ERROR] Shopify API connection failed: {e}")
        sys.exit(1)


def process_products(edges):
    active_products = {}
    skipped_count = 0
    
    for edge in edges:
        p = edge.get("node", {})
        title = p.get("title", "")
        handle = p.get("handle", "")
        pid = p.get("id", "")
        featured_image = p.get("featuredImage")
        variants_edges = p.get("variants", {}).get("edges", [])
        
        # Apply draft filters
        has_no_image = not featured_image or not featured_image.get("url")
        all_variants_zero_price = len(variants_edges) > 0 and all(
            float(v.get("node", {}).get("price", {}).get("amount", 0)) == 0 
            for v in variants_edges
        )
        has_draft_title = title and ("draft" in title.lower() or "test" in title.lower())
        
        if has_no_image or all_variants_zero_price or has_draft_title:
            print(f"  [SKIP DRAFT] Excluded draft product: {title} (handle: {handle})")
            skipped_count += 1
            continue
            
        variants = {}
        for v_edge in variants_edges:
            v = v_edge.get("node", {})
            vid = v.get("id", "")
            v_title = v.get("title", "")
            v_price = float(v.get("price", {}).get("amount", 0))
            variants[vid] = {
                "title": v_title,
                "price": v_price
            }
            
        active_products[pid] = {
            "title": title,
            "handle": handle,
            "variants": variants
        }
        
    print(f"Successfully processed {len(active_products)} active products (skipped {skipped_count} drafts).")
    return active_products


def load_existing_purchase_modes():
    modes = {}
    if os.path.exists(EXCEL_PATH):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            if "Products Summary" in wb.sheetnames:
                ws = wb["Products Summary"]
                for row in range(5, ws.max_row + 1):
                    name_val = ws.cell(row=row, column=1).value
                    mode_val = ws.cell(row=row, column=5).value
                    if name_val and mode_val:
                        modes[str(name_val).strip().lower()] = str(mode_val).strip()
            print(f"Loaded {len(modes)} product purchase modes from existing Excel sheet.")
        except Exception as e:
            print(f"[WARNING] Could not parse existing Excel sheet: {e}. Defaulting all to 'Sale'.")
    return modes


def get_dimensions_lookup():
    index_dims = {}
    
    # 1. Parse index.html product cards
    if os.path.exists(INDEX_HTML):
        with open(INDEX_HTML, 'r', encoding='utf-8') as f:
            index_html = f.read()
        dim_matches = re.findall(r'class="product-card"[^>]*data-handle="([^"]+)"[^>]*data-dim="([^"]+)"', index_html)
        for handle, dim in dim_matches:
            index_dims[handle] = dim.strip()
            
    # 2. Parse product.js dimensions
    js_dims = {}
    if os.path.exists(PRODUCT_JS):
        with open(PRODUCT_JS, 'r', encoding='utf-8') as f:
            product_js = f.read()
        handle_blocks = re.findall(r'"([^"]+)":\s*\{[^}]*?dimension:\s*"([^"]+)"', product_js, re.DOTALL)
        for handle, dim in handle_blocks:
            js_dims[handle] = dim.strip()
            
    # 3. Hardcoded lookup from searchIndex.json fallback
    search_index_dims = {
        "round-dining-chair": "50cm x 50cm x 77cm (Seat Height: 45cm)",
        "round-dining-table": "183cm x 107cm x 76cm",
        "poster-bed": "204cm x 220cm x 197cm (Seat Height: 48cm)",
        "dining-table": "198cm x 92cm x 75cm / 183cm x 92cm x 75cm",
        "study-chair": "60cm x 60cm x 84cm (Seat Height: 45cm)",
        "barstool-01": "48cm x 50cm x 100cm (Seat Height: 75cm)"
    }
    
    return index_dims, js_dims, search_index_dims


def resolve_dimension(handle, index_dims, js_dims, search_index_dims):
    return index_dims.get(handle) or js_dims.get(handle) or search_index_dims.get(handle) or "Not specified"


def generate_excel(products, existing_modes):
    wb = openpyxl.Workbook()
    
    # Sheet 1: Products Summary
    ws_summary = wb.active
    ws_summary.title = "Products Summary"
    
    # Fonts and styles
    header_fill = PatternFill(start_color="1A3636", end_color="1A3636", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1A3636")
    regular_font = Font(name="Calibri", size=11)
    italic_font = Font(name="Calibri", size=9, italic=True)
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    thick_bottom = Side(border_style="medium", color="1A3636")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)
    
    # Sheet title block
    ws_summary.append(["NOKU STUDIO — PRODUCT PORTFOLIO"])
    ws_summary.cell(row=1, column=1).font = title_font
    ws_summary.row_dimensions[1].height = 25
    ws_summary.append(["Generated from live Shopify catalog. All prices in INR."])
    ws_summary.cell(row=2, column=1).font = italic_font
    ws_summary.append([]) # Empty row
    
    # Headers
    summary_headers = ["Product Name", "Dimensions", "Price Range", "Number of Variants", "Purchase mode"]
    ws_summary.append(summary_headers)
    header_row_idx = 4
    ws_summary.row_dimensions[header_row_idx].height = 24
    
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx != 4 else "right", vertical="center")
        cell.border = border_header
        
    index_dims, js_dims, search_index_dims = get_dimensions_lookup()
    
    # Sort products alphabetically
    sorted_pids = sorted(products.keys(), key=lambda k: products[k]["title"])
    
    # Populate Summary Sheet
    row_idx = 5
    for pid in sorted_pids:
        p = products[pid]
        title = p["title"]
        handle = p["handle"]
        variants = p["variants"]
        
        dim = resolve_dimension(handle, index_dims, js_dims, search_index_dims)
        
        # Calculate price range
        prices = [v["price"] for v in variants.values() if v["price"] is not None]
        if prices:
            min_p = min(prices)
            max_p = max(prices)
            if min_p == max_p:
                price_range = "TBD" if min_p == 0 else f"₹{min_p:,.0f}"
            else:
                price_range = f"₹{min_p:,.0f} - ₹{max_p:,.0f}"
        else:
            price_range = "Not specified"
            
        num_variants = len(variants)
        
        # Retrieve or default Purchase Mode
        mode = existing_modes.get(title.strip().lower(), "Sale")
        
        ws_summary.append([title, dim, price_range, num_variants, mode])
        ws_summary.row_dimensions[row_idx].height = 20
        
        for col_idx in range(1, 6):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = border_all
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        row_idx += 1
        
    # Sheet 2: Product Variants Breakout
    ws_variants = wb.create_sheet(title="Product Variants")
    
    ws_variants.append(["NOKU STUDIO — DETAILED PRODUCT VARIANTS"])
    ws_variants.cell(row=1, column=1).font = title_font
    ws_variants.row_dimensions[1].height = 25
    ws_variants.append(["Breakdown of every available variant from store cache. All prices in INR."])
    ws_variants.cell(row=2, column=1).font = italic_font
    ws_variants.append([]) # Empty row
    
    variant_headers = ["Product Name", "Variant Name", "Dimensions", "Price"]
    ws_variants.append(variant_headers)
    v_header_row_idx = 4
    ws_variants.row_dimensions[v_header_row_idx].height = 24
    
    for col_idx, header in enumerate(variant_headers, start=1):
        cell = ws_variants.cell(row=v_header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx != 4 else "right", vertical="center")
        cell.border = border_header
        
    v_row_idx = 5
    for pid in sorted_pids:
        p = products[pid]
        title = p["title"]
        handle = p["handle"]
        variants = p["variants"]
        dim = resolve_dimension(handle, index_dims, js_dims, search_index_dims)
        
        if not variants:
            ws_variants.append([title, "Base Style", dim, 0])
            ws_variants.row_dimensions[v_row_idx].height = 18
            for col_idx in range(1, 5):
                cell = ws_variants.cell(row=v_row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = border_all
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.value = "TBD"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            v_row_idx += 1
        else:
            for vid, v in variants.items():
                v_title = v["title"]
                v_price = v["price"]
                
                ws_variants.append([title, v_title, dim, v_price])
                ws_variants.row_dimensions[v_row_idx].height = 18
                for col_idx in range(1, 5):
                    cell = ws_variants.cell(row=v_row_idx, column=col_idx)
                    cell.font = regular_font
                    cell.border = border_all
                    if col_idx == 4:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if v_price == 0:
                            cell.value = "TBD"
                        else:
                            cell.number_format = '"₹"#,##0'
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                v_row_idx += 1
                
    # Auto-adjust column widths
    for ws in [ws_summary, ws_variants]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 3 and cell.value:
                    val_str = str(cell.value)
                    if isinstance(cell.value, (int, float)) and cell.number_format and "₹" in cell.number_format:
                        val_str = f"₹{cell.value:,.0f}"
                    max_len = max(max_len, len(val_str))
            adjusted_width = max(max_len + 4, 12)
            if col_letter in ['A', 'B', 'C']:
                adjusted_width = min(adjusted_width, 45)
            ws.column_dimensions[col_letter].width = adjusted_width
            
    wb.save(EXCEL_PATH)
    print(f"Excel portfolio updated successfully at: {EXCEL_PATH}")


def resolve_display_data(excel_products, product_cache):
    title_to_product = {}
    for gid, pdata in product_cache.items():
        title = pdata.get("title", "").strip().lower()
        title_to_product[title] = (gid, pdata)

    display_handles     = []
    display_titles      = []
    display_variant_ids = []

    for name, mode in excel_products.items():
        name_lower = name.strip().lower()

        if name_lower not in title_to_product:
            continue

        gid, pdata = title_to_product[name_lower]
        handle = pdata.get("handle", "")
        variants = pdata.get("variants", {})

        if mode.strip().lower() == "display":
            if handle and handle not in display_handles:
                display_handles.append(handle)

            raw_title = name_lower
            if raw_title not in display_titles:
                display_titles.append(raw_title)
            hyphenated = raw_title.replace(" ", "-")
            if hyphenated != raw_title and hyphenated not in display_titles:
                display_titles.append(hyphenated)

            for v_gid in variants.keys():
                numeric_id = v_gid.split("/")[-1]
                if numeric_id and numeric_id not in display_variant_ids:
                    display_variant_ids.append(numeric_id)

    return display_handles, display_titles, display_variant_ids


def format_handles_array(handles, indent="  "):
    if not handles:
        return "[]"
    items = [f"{indent}'{h}'" for h in handles]
    return "[\n" + ",\n".join(items) + "\n]"


def format_titles_array(titles, indent="    "):
    if not titles:
        return "[]"
    items = [f'{indent}"{t}"' for t in titles]
    return "[\n" + ",\n".join(items) + "\n  ]"


def format_variant_ids_array(ids, indent="    ", chunk_size=10):
    if not ids:
        return "[]"
    lines = []
    for i in range(0, len(ids), chunk_size):
        chunk = ids[i:i + chunk_size]
        quoted = [f'"{vid}"' for vid in chunk]
        lines.append(f'{indent}' + ", ".join(quoted))
    return "[\n" + ",\n".join(lines) + "\n  ]"


def update_js_file(filepath, display_handles, display_titles, display_variant_ids):
    if not os.path.exists(filepath):
        print(f"[ERROR] JS file not found: {filepath}")
        return False

    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    original_content = content

    # 1. Replace top-level DISPLAY_ONLY_HANDLES
    new_handles_str = format_handles_array(display_handles)
    handles_pattern = re.compile(
        r"(// Display-only products list.*?\n)"
        r"(const DISPLAY_ONLY_HANDLES\s*=\s*)\[.*?\];",
        re.DOTALL
    )
    content, _ = re.subn(handles_pattern, rf"\g<1>\g<2>{new_handles_str};", content)

    # 2. Replace displayTitles inside isDisplayItem()
    new_titles_str = format_titles_array(display_titles)
    titles_pattern = re.compile(
        r"(const displayTitles\s*=\s*)\[.*?\];",
        re.DOTALL
    )
    content, _ = re.subn(titles_pattern, rf"\g<1>{new_titles_str};", content)

    # 3. Replace displayVariantIds
    new_vids_str = format_variant_ids_array(display_variant_ids)
    vids_pattern = re.compile(
        r"(const displayVariantIds\s*=\s*)\[.*?\];",
        re.DOTALL
    )
    content, _ = re.subn(vids_pattern, rf"\g<1>{new_vids_str};", content)

    # 4. Replace displayHandles inside updateCartUI()
    new_dh_str = format_handles_array(display_handles, indent="    ")
    dh_pattern = re.compile(
        r"(const displayHandles\s*=\s*)\[.*?\];",
        re.DOTALL
    )
    content, _ = re.subn(dh_pattern, rf"\g<1>{new_dh_str};", content)

    changed = content != original_content
    if changed:
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"  [SAVED] {os.path.basename(filepath)} updated.")
    else:
        print(f"  [OK] {os.path.basename(filepath)} is already up to date.")
    return changed


def update_gitignore():
    if os.path.exists(GITIGNORE_PATH):
        with open(GITIGNORE_PATH, 'r', encoding='utf-8') as f:
            content = f.read()
        if "noku_products.xlsx" not in content:
            new_content = content
            if not content.endswith('\n'):
                new_content += '\n'
            new_content += "# Generated Product Excel Spreadsheet\nnoku_products.xlsx\n"
            with open(GITIGNORE_PATH, 'w', encoding='utf-8') as f:
                f.write(new_content)
            print("Updated .gitignore to exclude noku_products.xlsx")


def main():
    print("=" * 60)
    print("  NOKU STUDIO — UNIFIED SHOPIFY WEBSITE UPDATE SCRIPT")
    print("=" * 60)
    
    # Step 1: Load Credentials
    print("\n[1/5] Loading Shopify Storefront Credentials...")
    config = load_shopify_config()
    print(f"      Store Domain: {config['shopDomain']}")
    print(f"      API Version:  {config['apiVersion']}")
    
    # Step 2: Fetch Live Catalog
    print("\n[2/5] Fetching live products from Shopify storefront API...")
    edges = fetch_products(config)
    products = process_products(edges)
    
    # Write to product_cache.json
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(products, f, indent=2)
    print(f"      Updated catalog cache: {CACHE_PATH}")
    
    # Step 3: Load existing purchase modes and generate Excel Portfolio
    print("\n[3/5] Syncing Excel Portfolio...")
    existing_modes = load_existing_purchase_modes()
    generate_excel(products, existing_modes)
    update_gitignore()
    
    # Step 4: Parse final purchase modes from Excel on disk
    print("\n[4/5] Loading final purchase modes from spreadsheet...")
    excel_products = {}
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "Products Summary" in wb.sheetnames:
        ws = wb["Products Summary"]
        for row in range(5, ws.max_row + 1):
            name_val = ws.cell(row=row, column=1).value
            mode_val = ws.cell(row=row, column=5).value
            if name_val and mode_val:
                excel_products[str(name_val).strip()] = str(mode_val).strip()
    print(f"      Loaded {len(excel_products)} product purchase mode records.")
    
    # Step 5: Resolve Display Data and Update JavaScript configs
    print("\n[5/5] Syncing display settings with Javascript files...")
    display_handles, display_titles, display_variant_ids = resolve_display_data(excel_products, products)
    
    print(f"      Resolved display-only configurations:")
    print(f"        - Handles: {len(display_handles)}")
    print(f"        - Titles:  {len(display_titles)}")
    print(f"        - Variant IDs: {len(display_variant_ids)}")
    
    js_files = [PRODUCT_JS, PRODUCTS_JS, SHOPIFY_JS]
    any_changed = False
    for filepath in js_files:
        if update_js_file(filepath, display_handles, display_titles, display_variant_ids):
            any_changed = True
            
    print("\n" + "=" * 60)
    if any_changed:
        print("  [SUCCESS] Website updated! JS configurations synced with Shopify.")
    else:
        print("  [SUCCESS] Sync complete! No Javascript files required updates.")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
